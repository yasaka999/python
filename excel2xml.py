import openpyxl
wb = openpyxl.load_workbook("../files/hlj_demo.xlsx", data_only=True)
ws = wb.worksheets[0]
# 读取
# print(sh.rows,type(sh.rows))
# 按行读取，每一行的数据放到一个元组中

list_tuple=[]
    # 逐行循环读取excel的数据
for value in ws.values:
    list_tuple.append(value)
#print(list_tuple,type(list_tuple))
print (ws.max_row)
print (ws.max_column)
print (list_tuple[0][1])
for i in range(len(list_tuple)) :
    if list_tuple[i][3]=="单集" :
        print (list_tuple[i])
