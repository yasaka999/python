from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# 处理表格
#old_filename = '订购明细' + str(b_yesterday) + '.xlsx'
# wb = load_workbook(old_filename,data_only=True)
wb = load_workbook('LT-order-detail-2021-07-08.xlsx')
# 处理日订购表
ws0 = wb.worksheets[0]
ws2 = wb.worksheets[2]

nrows0 = ws0.max_row  # 获得最大行数
ncols0 = ws0.max_column # 获取最大列数
nrows2 = ws2.max_row
ws0.cell(row=nrows0,column=25).value= ws0.cell(row=nrows0,column=24).value
ws0.cell(row=nrows0,column=26).value = (ws2.cell(row=nrows2-1,column=8).value - ws2.cell(row=nrows2,column=9).value) * 9.9
ws0.cell(row=nrows0,column=27).value = ws0.cell(row=nrows0,column=25).value + ws0.cell(row=nrows0,column=26).value

#格式化
thin = Side(border_style="thin", color="000000")#边框样式，颜色
border = Border(left=thin, right=thin, top=thin, bottom=thin)#边框的位置
font = Font(size=10, bold=False, name='微软雅黑',  color="000000")#字体大小，加粗，字体名称，字体名字

for i in range (25,28):
	ws0.cell(row=nrows0,column=i).alignment = Alignment(horizontal='center', vertical='center')
	ws0.cell(row=nrows0,column=i).border = border
	ws0.cell(row=nrows0,column=i).font = font
	ws0.cell(row=nrows0,column=i).fill = PatternFill('solid', fgColor="FFA500")

wb.save('test.xlsx')