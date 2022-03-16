# 先增加数据
# 格式化
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side



wb = load_workbook('订购明细6.6日数据.xlsx',data_only=True)
ws0 = wb.worksheets[0]
#ws.title = "故障1"
today = datetime.date.today()
product_1_AAA = 34
product_1_C3 = 35
product_2_AAA = 3
product_2_C3 = 2
# product_3_AAA = 
nrows = ws0.max_row
d_1 = "=B"+str(nrows+1)+"-C"+str(nrows+1)

ws0.append([today,product_1_AAA,product_1_C3,d_1,product_2_AAA,product_2_C3,product_2_C3-product_2_AAA])

ws1 = wb.worksheets[1]
ws1.append([today,product_1_AAA,product_1_C3,product_1_C3-product_1_AAA,product_2_AAA,product_2_C3,product_2_C3-product_2_AAA])
#格式化
thin = Side(border_style="thin", color="000000")#边框样式，颜色
border = Border(left=thin, right=thin, top=thin, bottom=thin)#边框的位置
font = Font(size=10, bold=False, name='微软雅黑',  color="000000")#字体大小，加粗，字体名称，字体名字

nrows = ws0.max_row  # 获得行数
ncols = ws0.max_column
for i in range(nrows-1,nrows):
    for j in range(ncols):
        ws0.cell(row=i+1, column=j+1).alignment = Alignment(horizontal='center', vertical='center')
        ws0.cell(row=i+1, column=j+1).border = border
        ws0.cell(row=i+1, column=j+1).font = font
# 有差异的标成红色
for a in range(nrows-1,nrows):
	for b in (4,7):
		if ws0.cell(row=a+1, column=b).value != 0 :
			ws0.cell(row=a+1, column=b).font = Font(size=10, bold=True, name='微软雅黑',  color="FF0000")

ws0.cell(nrows,1).number_format = 'm/d'
new_filename = '订购明细'+str(today)+'.xlsx'
wb.save(new_filename)
