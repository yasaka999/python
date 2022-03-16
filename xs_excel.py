#from openpyxl import Workbook
from openpyxl import load_workbook
import re

#处理表格

wb = load_workbook('违规低俗下线节目汇总.xlsx',data_only=True)

org_name=[]
out_name1=[]
out_name2=[]
file1= open('书名号.txt','w')
file2= open('书名号-异常.txt','w')
file3= open('全名.txt','w')
file4= open('全名-带括号.txt','w')
file5= open('待比对.txt','w')
# 读取所有名称
for i in range(3):
        ws = wb.worksheets[i]
        nrows = ws.max_row  # 获得最大行数
        for j in range(2,nrows+1):
                org_name.append(ws.cell(row=j,column=4).value)
# 按是否包含书名号拆成两个名称列表
for j in range(len(org_name)):
        if  ("》" in org_name[j]) or ("《" in org_name[j]) :
                out_name1.append(org_name[j])
        else:
                out_name2.append(org_name[j])
# 处理书名号，提取名称
for m in range(len(out_name1)):
        p1 = re.compile(r'[《](.*?)[》]', re.S) 
        arr = re.findall(p1, out_name1[m])
        if arr:
#                continue
#                print (out_name1[m],arr)
                for i in arr:
                        print (out_name1[m]+'  |  ',i,file=file1)
                        print (i,file=file5)

        else:
                print (out_name1[m],file=file2)
# 处理全名的带括号
for n in range(len(out_name2)):
        p1 = re.compile(r'(.*?)[（]', re.S) 
        arr = re.findall(p1, out_name2[n])
        if arr:
#                continue
#                print (out_name1[m],arr)
                print (out_name2[n]+'  |  ',arr[0],file=file4)
                print (arr[0],file=file5)

        else:
                print (out_name2[n],file=file3)
                print (out_name2[n],file=file5)

file1.close
file2.close
file3.close
file4.close
file5.close

