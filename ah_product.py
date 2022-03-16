import cx_Oracle
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import urllib.request 
from http import cookiejar    #保存cookie用的
from urllib import parse 
import json

#获取aaa的订购退订数据
connection = cx_Oracle.connect("wacos", "AHsop2020","10.2.13.82:1521/orcl")
cursor = connection.cursor()
cursor.execute("""
select ee.days,
       nvl(dg_5, 0),
       nvl(td_jc1, 0),
       nvl(dg_6, 0),
       nvl(td_jc3, 0),
       nvl(dg_3, 0),
       nvl(td_lx2, 0),
       nvl(dg_1, 0),
       nvl(td_jc2, 0),
       nvl(dg_4, 0),
       nvl(td_jc4, 0),
       nvl(dg_9, 0),
       nvl(td_lx5, 0),
       nvl(dg_10, 0),
       nvl(td_lx6, 0),
       nvl(dg_11, 0),
       nvl(td_lx7, 0),
       nvl(dg_12, 0),
       nvl(td_lx8, 0),
       nvl(dg_7, 0),
       nvl(td_jc5, 0),
       nvl((nvl(dg_2, 0) + nvl(dg_8, 0)), 0) as "other_order",
       nvl((nvl(td_jc8, 0) + nvl(td_jc6, 0) + nvl(td_jc7, 0)), 0) as "other_disorser",
       nvl((nvl(dg_5, 0) + nvl(dg_6, 0) + nvl(dg_3, 0) + nvl(dg_1, 0) +
           nvl(dg_4, 0) + nvl(dg_2, 0) + nvl(dg_7, 0) + nvl(dg_8, 0) +
           nvl(dg_10, 0) + nvl(dg_11, 0) + nvl(dg_12, 0) + nvl(dg_4, 9)),
           0) as "order_sum",
       nvl((td_jc1 + td_jc3 + nvl(td_lx2, 0) + td_jc2 + td_jc4 + td_jc5 +
           td_jc8 + td_jc6 + td_jc7 + nvl(td_lx5, 0) + nvl(td_lx6, 0) +
           nvl(td_lx7, 0) + nvl(td_lx8, 0)),
           0) as "disorder_sum"
  from (select *
          from (select a.date1,
                       max(case package1
                             when 21 then
                              sum1
                             else
                              0
                           end) td_lx1,
                       max(case package1
                             when 23 then
                              sum1
                             else
                              0
                           end) td_lx2,
                       max(case package1
                             when 16 then
                              sum1
                             else
                              0
                           end) td_lx3,
                       max(case package1
                             when 5 then
                              sum1
                             else
                              0
                           end) td_lx4,
                       max(case package1
                             when 32 then
                              sum1
                             else
                              0
                           end) td_lx5,
                       max(case package1
                             when 33 then
                              sum1
                             else
                              0
                           end) td_lx6,
                       max(case package1
                             when 34 then
                              sum1
                             else
                              0
                           end) td_lx7,
                       max(case package1
                             when 35 then
                              sum1
                             else
                              0
                           end) td_lx8
                  from (select to_char(i.completedate, 'yyyy/mm/dd') date1,
                               i.iptvproductid package1,
                               count(1) sum1
                          from iptvproductorder i
                         where i.optype = 3
                              --  and i.completedate > to_date('2020/9/1', 'yyyy/mm/dd')
                           and to_char(i.completedate, 'yyyy/mm/dd') <
                               to_char(sysdate, 'yyyy/mm/dd')
                           and i.iptvproductid in
                               (select p.iptvproductid
                                  from iptvproduct p
                                 where p.status = 1
                                   and nvl(p.chargingtype, 99) != 2)
                         group by to_char(i.completedate, 'yyyy/mm/dd'),
                                  i.iptvproductid
                        
                        ) a
                 group by a.date1) aa
          full join (select to_char((trunc(sysdate - 1000) + rownum),
                                   'yyyy/mm/dd') as days
                      from dual
                    connect by rownum <= 1000) dd
            on aa.date1 = dd.days) ee
  full join (select *
               from (select b.date2,
                            max(case package1
                                  when 31 then
                                   sum1
                                  else
                                   0
                                end) td_jc1,
                            max(case package1
                                  when 20 then
                                   sum1
                                  else
                                   0
                                end) td_jc2,
                            max(case package1
                                  when 22 then
                                   sum1
                                  else
                                   0
                                end) td_jc3,
                            max(case package1
                                  when 24 then
                                   sum1
                                  else
                                   0
                                end) td_jc4,
                            max(case package1
                                  when 25 then
                                   sum1
                                  else
                                   0
                                end) td_jc5,
                            max(case package1
                                  when 30 then
                                   sum1
                                  else
                                   0
                                end) td_jc6,
                            max(case package1
                                  when 26 then
                                   sum1
                                  else
                                   0
                                end) td_jc7,
                            max(case package1
                                  when 18 then
                                   sum1
                                  else
                                   0
                                end) td_jc8
                       from (
                             
                             select to_char(i.completedate, 'yyyy/mm/dd') date2,
                                     i.iptvproductid package1,
                                     count(1) sum1
                               from iptvproductorder i
                              where i.optype = 2
                                   --  and i.completedate > to_date('2020/9/1', 'yyyy/mm/dd')
                                and to_char(i.completedate, 'yyyy/mm/dd') <
                                    to_char(sysdate, 'yyyy/mm/dd')
                                and i.iptvproductid in
                                    (select p.iptvproductid
                                       from iptvproduct p
                                      where p.status = 1
                                        and p.chargingtype = 2)
                              group by to_char(i.completedate, 'yyyy/mm/dd'),
                                        i.iptvproductid
                             
                             ) b
                      group by b.date2) bb
               full join (select to_char((trunc(sysdate - 1000) + rownum),
                                        'yyyy/mm/dd') as days
                           from dual
                         connect by rownum <= 1000) dd
                 on bb.date2 = dd.days) ff

    on ee.days = ff.days
  full join (select *
               from (select date3,
                            max(case package1
                                  when 20 then
                                   sum1
                                  else
                                   0
                                end) as dg_1,
                            max(case package1
                                  when 30 then
                                   sum1
                                  else
                                   0
                                end) dg_2,
                            max(case package1
                                  when 23 then
                                   sum1
                                  else
                                   0
                                end) dg_3,
                            max(case package1
                                  when 24 then
                                   sum1
                                  else
                                   0
                                end) dg_4,
                            max(case package1
                                  when 31 then
                                   sum1
                                  else
                                   0
                                end) dg_5,
                            max(case package1
                                  when 22 then
                                   sum1
                                  else
                                   0
                                end) dg_6,
                            max(case package1
                                  when 25 then
                                   sum1
                                  else
                                   0
                                end) dg_7,
                            max(case package1
                                  when 26 then
                                   sum1
                                  else
                                   0
                                end) dg_8,
                            max(case package1
                                  when 32 then
                                   sum1
                                  else
                                   0
                                end) dg_9,
                            max(case package1
                                  when 33 then
                                   sum1
                                  else
                                   0
                                end) dg_10,
                            max(case package1
                                  when 34 then
                                   sum1
                                  else
                                   0
                                end) dg_11,
                            max(case package1
                                  when 35 then
                                   sum1
                                  else
                                   0
                                end) dg_12
                       from (select to_char(i.operatetime, 'yyyy/mm/dd') date3,
                                    i.iptvproductid package1,
                                    count(1) sum1
                               from iptvproductorder i
                              where i.optype = 2
                             --and i.operatetime > to_date('2020/9/1','yyyy/mm/dd')
                              group by to_char(i.operatetime, 'yyyy/mm/dd'),
                                       i.iptvproductid) a
                      group by date3) cc
               full join (select to_char((trunc(sysdate - 1000) + rownum),
                                        'yyyy/mm/dd') as days
                           from dual
                         connect by rownum <= 1000) dd
                 on cc.date3 = dd.days) gg
    on gg.days = ff.days
 order by ee.days desc

""")
rows=cursor.fetchmany(2) 
cursor.close()
connection.close()

aaa = rows[1]


#请求bi web 获得c3数据


login_url = "http://10.178.50.164:8080/bi-portal/login/login.do"

login_form_data = {
    "name": "reportadmin",
    "password": "Utiptv@2021",
}
 #1.3 发送登录请求POST
cook_jar = cookiejar.CookieJar()
#定义有添加 cook 功能的 处理器
cook_hanlder = urllib.request.HTTPCookieProcessor(cook_jar)
#根据处理器 生成 opener
opener = urllib.request.build_opener(cook_hanlder)

#带着参数 发送post请求
#1 参数 将来 需要转译 转码； 2 post 请求的 data 要求是bytes
login_str = parse.urlencode(login_form_data).encode('utf-8')
login_request = urllib.request.Request(login_url,data=login_str)

#如果登录成功，cookjar自动保存cookie
opener.open(login_request)
token_url = "http://10.178.50.164:8080/bi-portal/configure/getToken.do"
#token_request = urllib.request.Request(token_url)
response = opener.open(token_url)
#bytes -->str
data = json.loads(response.read())
token = data['token']

# 开始请求，获取数据存成字典c3_order
yesterday = (datetime.datetime.now()+datetime.timedelta(days=-1)).strftime('%Y%m%d')

query_url = "http://10.178.50.164:8080/statplatform5/query"

q_data1="?qdi=queryc3ordersinglepoint&starttime="+yesterday+"&endtime="+yesterday+"&datetype=D&token="+token+"&sysid=t&userid=test1"
url = query_url+q_data1

q_request = urllib.request.Request(url)

res = urllib.request.urlopen(q_request)
result = json.loads(res.read())
c3_order = {}
for i in range(len(result["subject"])):
    c3_order[result["subject"][i]["CODE"]] = result["subject"][i]["ORDERCOUNT"]
#    print (result["subject"][i]["NAME"],result["subject"][i]["ORDERCOUNT"],result["subject"][i]["CODE"])

q_data2="?qdi=queryc3ordermonthpoint&starttime="+yesterday+"&endtime="+yesterday+"&datetype=D&type=0%27%2C%273%27%2C%274%27%2C%275%27%2C%27&token="+token+"&sysid=t&userid=test1"
url = query_url+q_data2

q_request = urllib.request.Request(url)

res = urllib.request.urlopen(q_request)
result = json.loads(res.read())

for i in range(len(result["subject"])):
    c3_order[result["subject"][i]["CODE"]] = result["subject"][i]["NEW_ORDERCOUNT"]
#    print (result["subject"][i]["NAME"],result["subject"][i]["NEW_ORDERCOUNT"],result["subject"][i]["CODE"])
print (c3_order)



today = datetime.date.today()
yesterday = datetime.date.today() + datetime.timedelta(-1)



# aaa数据赋值
(p1_aaa,t1,p2_aaa,t2,p3_aaa,t3,p4_aaa,t4,p5_aaa,t5,p6_aaa,t6,p7_aaa,t7,p8_aaa,t8,p9_aaa,t9)\
= (aaa[1],aaa[2],aaa[19],aaa[20],aaa[3],aaa[4],aaa[5],aaa[6],aaa[9],aaa[10],aaa[11],aaa[12],\
        aaa[13],aaa[14],aaa[15],aaa[16],aaa[17],aaa[18])
# 新开赠送
p3_add = 0

# c3数据赋值
p1_c3 = c3_order.get("2400057406@213", 0)
p2_c3 = c3_order.get("2400057404@213", 0)
p3_c3 = c3_order.get("2400057401@213", 0)
p4_c3 = c3_order.get("2400057402@213", 0)
p5_c3 = c3_order.get("2400057403@213", 0)
p6_c3 = c3_order.get("2400057407@213", 0)
p7_c3 = c3_order.get("2400057408@213", 0)
p8_c3 = c3_order.get("2400057409@213", 0)
p9_c3 = c3_order.get("2400057410@213", 0)

#差异
d_1 = p1_aaa - p1_c3
d_2 = p2_aaa - p2_c3
d_3 = p3_aaa - p3_c3 + p3_add
d_4 = p4_aaa - p4_c3
d_5 = p5_aaa - p5_c3
d_6 = p6_aaa - p6_c3
d_7 = p7_aaa - p7_c3
d_8 = p8_aaa - p8_c3
d_9 = p9_aaa - p9_c3
#汇总
total_order_v = p1_aaa + p3_aaa - p3_add + p4_aaa + p5_aaa
total_t_v = t1 + t3 + t4 + t5
total_sum_v = p1_aaa * 5.0 + p2_aaa * 5.0 + (p3_aaa - p3_add) * 15 + p4_aaa * 9.9 + p5_aaa * 99
total_order_c = p6_aaa + p7_aaa + p8_aaa + p9_aaa
total_t_c = t6 + t7 + t8 + t9
total_sum_c = p6_aaa * 2.9 + p7_aaa * 1.9 + p8_aaa * 1.9 + p9_aaa * 9.9

#处理表格

wb = load_workbook('订购明细2021.6.8.xlsx',data_only=True)
#wb = load_workbook('订购明细6.6日数据.xlsx')
#处理日订购表
ws0 = wb.worksheets[0]

ws0.append([yesterday,p1_aaa,p1_c3,d_1,p2_aaa,p2_c3,d_2,p3_aaa,p3_c3,p3_add,d_3,\
        p4_aaa,p4_c3,d_4,p5_aaa,p5_c3,d_5,total_order_v,t1,t3,t4,t5,total_t_v,total_sum_v])



#格式化
thin = Side(border_style="thin", color="000000")#边框样式，颜色
border = Border(left=thin, right=thin, top=thin, bottom=thin)#边框的位置
font = Font(size=10, bold=False, name='微软雅黑',  color="000000")#字体大小，加粗，字体名称，字体名字

nrows = ws0.max_row  # 获得最大行数
ncols = ws0.max_column # 获取最大列数
for i in range(nrows-1,nrows):
    for j in range(ncols):
        ws0.cell(row=i+1, column=j+1).alignment = Alignment(horizontal='center', vertical='center')
        ws0.cell(row=i+1, column=j+1).border = border
        ws0.cell(row=i+1, column=j+1).font = font
# 有差异的标成红色
for a in range(nrows-1,nrows):
        for b in (4,7,11,14,17):
                ws0.cell(row=a+1, column=b).fill = PatternFill('solid', fgColor="C0C0C0")
                if ws0.cell(row=a+1, column=b).value != 0 :
                        ws0.cell(row=a+1, column=b).font = Font(size=10, bold=True, name='微软雅黑',  color="FF0000")

ws0.cell(nrows,1).number_format = 'm/d'
ws0.cell(nrows,24).font = Font (color="FF0000")

#处理频道订购表
ws1 = wb.worksheets[1]
ws1.append([yesterday,p6_aaa,p6_c3,d_6,p7_aaa,p7_c3,d_7,p8_aaa,p8_c3,d_8,p9_aaa,p9_c3,d_9,\
        total_order_c,t6,t7,t8,t9,total_t_c,total_sum_c])


nrows = ws1.max_row  # 获得最大行数
ncols = ws1.max_column # 获取最大列数
for i in range(nrows-1,nrows):
    for j in range(ncols):
        ws1.cell(row=i+1, column=j+1).alignment = Alignment(horizontal='center', vertical='center')
        ws1.cell(row=i+1, column=j+1).border = border
        ws1.cell(row=i+1, column=j+1).font = font
# 有差异的标成红色
for a in range(nrows-1,nrows):
        for b in (4,7,10,13):
                ws1.cell(row=a+1, column=b).fill = PatternFill('solid', fgColor="C0C0C0")
                if ws1.cell(row=a+1, column=b).value != 0 :
                        ws1.cell(row=a+1, column=b).font = Font(size=10, bold=True, name='微软雅黑',  color="FF0000")

ws1.cell(nrows,1).number_format = 'm/d'
ws1.cell(nrows,20).font = Font (color="FF0000")


new_filename = '订购明细'+str(yesterday)+'.xlsx'
wb.save(new_filename)