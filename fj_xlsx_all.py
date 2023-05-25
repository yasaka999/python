import sys
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# 获取命令行参数
input_file = sys.argv[1]

# 读取原始xlsx文件的所有sheet
xls = pd.ExcelFile(input_file)

# 创建新的Excel文件
output_file = input_file.replace('.xlsx', '_new.xlsx')
writer = pd.ExcelWriter(output_file)

# 遍历原始文件中的每个sheet
for sheet_name in xls.sheet_names:
    # 读取当前sheet的数据
    df = pd.read_excel(input_file, sheet_name=sheet_name)
    if sheet_name == '用户运营数据':
        # 指定要保留的指标名称
        selected_indicators = ['全网用户数', '新增用户', '活跃用户','活跃率','新增用户占活跃比例','沉默用户']

        # 使用条件筛选保留对应指标名称的行
        df_selected = df[df['指标名称'].isin(selected_indicators)]

        df_selected.to_excel(writer, sheet_name=sheet_name, index=False)

    elif sheet_name =='直播分析':
        # 指定要保留的指标名称
        selected_indicators = ['收视人数', '收视次数', '收视时长(h)','人均收视时长(h)']

        # 使用条件筛选保留对应指标名称的行
        df_selected = df[df['指标名称'].isin(selected_indicators)]
        # 特定数据取整
        df_selected.iloc[2,1]= round(float(df_selected.iloc[2,1]))

        df_selected.to_excel(writer, sheet_name=sheet_name, index=False)
        
    elif sheet_name == '点播分析':
        # 指定要保留的指标名称
        selected_indicators = ['收视人数', '收视次数', '收视时长(h)','次均收视时长(h)','人均收视时长(h)']

        # 使用条件筛选保留对应指标名称的行
        df_selected = df[df['指标名称'].isin(selected_indicators)]
        # 特定数据取整
        df_selected.iloc[2,1]= round(float(df_selected.iloc[2,1]))

        df_selected.to_excel(writer, sheet_name=sheet_name, index=False)        
    
    elif sheet_name == '点播类型分析':
        df[df.columns[1]] = df[df.columns[1]].replace('%', '', regex=True).astype(float)
        df_selected = df.sort_values(by=df.columns[1], ascending=False)
        df_selected[df.columns[1]] = df_selected[df.columns[1]].astype(str) + '%'
        df_selected.to_excel(writer, sheet_name=sheet_name, index=False) 

    elif sheet_name == '播放内容top':
        # 定义函数，根据通用规则处理名称
        def process_channel_name(channel_name):
            # processed_name = channel_name.strip()  # 去除空格
            processed_name = re.sub(r'-第(.*)期.*', '', channel_name)
            processed_name = re.sub(r'-\d+-娱乐', '', processed_name)
            return processed_name

        # 处理名称
        df['节目收视次数TOP50'] = df['节目收视次数TOP50'].apply(process_channel_name)
        # 收视时长取整
        df['收视时长（h）'] = df['收视时长（h）'].round()

        # 处理百分比的列
        percent_columns = ['占比']
        df[percent_columns] = df[percent_columns].replace('%', '', regex=True).astype(float)

        # 选择需要保留原始内容的列
        preserve_columns = ['日期','分类']

        # 将这些列加入到分组的列中
        group_columns = ['节目收视次数TOP50'] + preserve_columns

        # 对选择的列进行分组和聚合求和操作
        df_grouped = df.groupby(group_columns).sum().reset_index()

        # 按收视时长倒序排列
        df_merged = df_grouped.groupby('分类').apply(lambda x: x.sort_values(by='收视时长（h）', ascending=False)).reset_index(drop=True)

        # 选择需要输出的列
        selected_columns = ['日期','分类','节目收视次数TOP50','收视时长（h）','收视次数','占比','收视人数']

        # 选择输出选定列的某些内容
        df_selected = df_merged[df_merged['分类'].isin(['电视剧', '电影', '综艺', '少儿'])][selected_columns]

        # 占比加上%
        df_selected['占比'] = df_selected['占比'].astype(str) + '%'

        # 保留每类的前10条记录
        df_selected = df_selected.groupby('分类').head(10)

        # 将处理后的数据写入新文件的对应sheet
        df_selected.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
# 保存新文件并关闭ExcelWriter,如果在win下有报错，用writer._save()
writer._save()
writer.close()

# 调整生成后的文件格式
workbook = load_workbook(output_file)
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                value = str(cell.value)
                if len(value) > max_length:
                    max_length = len(value.encode('utf-8'))  # 使用UTF-8编码计算字符长度
            except:
                pass
        adjusted_width = (max_length) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width

    # 设置单元格居中对齐
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

# 保存调整后的文件
workbook.save(output_file)

print(f"处理结果已保存至 {output_file}")
