import sys
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# 获取命令行参数
input_file = sys.argv[1]

# 读取原始xlsx文件的所有sheet
xls = pd.ExcelFile(input_file)

# 创建新的Excel文件
output_file = input_file.replace('.xlsx', '_new.xlsx')
writer = pd.ExcelWriter(output_file)

# 遍历原始文件中的每个sheet
for sheet_name in xls.sheet_names:
    # 读取当前sheet的数据
    df = pd.read_excel(input_file, sheet_name=sheet_name)

    # 定义函数，根据通用规则处理频道名称
    def process_channel_name(channel_name):
        processed_name = channel_name.strip()  # 去除空格
        processed_name = processed_name = re.sub(r'\(高清\)|高清|标清|\(标清\)|\（高清\）', '', channel_name)  # 去掉标高清
        return processed_name

    # 处理频道名称
    df['频道名称'] = df['频道名称'].apply(process_channel_name)

    # 将收视人数（UV）、收视次数（PV）和收视时长转换为数字格式
    numeric_columns = ['收视人数（UV）', '收视次数（PV）', '收视时长']
    df[numeric_columns] = df[numeric_columns].replace(',', '', regex=True).astype(float)

    # 收视时长取整
    df['收视时长'] = df['收视时长'].round()

    # 按频道名称合并相同频道的列，并对其他列的值进行求和
    df_merged = df.groupby('频道名称').sum().reset_index()

    # 按收视时长倒序排列
    df_merged = df_merged.sort_values(by='收视时长', ascending=False)

     # 选择需要输出的列
    selected_columns = ['频道名称','收视次数（PV）','收视人数（UV）', '收视时长']

    # 从合并后的DataFrame中筛选出选定的列
    df_selected = df_merged[selected_columns]

    # 保留前20条记录
    df_selected = df_selected.head(20)

    # 将处理后的数据写入新文件的对应sheet
    df_selected.to_excel(writer, sheet_name=sheet_name, index=False)

# 保存新文件并关闭ExcelWriter,如果在win下有报错，用writer._save()
writer.save()
writer.close()

# 调整生成后的文件格式
workbook = load_workbook(output_file)
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                value = str(cell.value)
                if len(value) > max_length:
                    max_length = len(value.encode('utf-8'))  # 使用UTF-8编码计算字符长度
            except:
                pass
        adjusted_width = (max_length) * 1.1
        sheet.column_dimensions[column_letter].width = adjusted_width

    # 设置单元格居中对齐
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

# 保存调整后的文件
workbook.save(output_file)

print(f"处理结果已保存至 {output_file}")
