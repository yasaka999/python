# 规则：
# 抽取节目名称规则，如下：
# 1、节目名称中带《》的，取《》内文字做为节目名，进行后续处理。
# 2、除第1种情况的节目，其他节目名称中带（），节目名取（前文字做为节目名，进行后续处理。
# 3、除第1、2种情况的节目，采用节目名称全部文字进行后续处理。
import re
import time
import cx_Oracle
import openpyxl

cx_Oracle.init_oracle_client(lib_dir="/Users/hanxiong/Downloads/instantclient_19_8")
# 处理表格

wb = openpyxl.load_workbook("../files/违规低俗下线节目汇总.xlsx", data_only=True)

org_name = []
out_name1 = []
out_name2 = []
file1 = open("../files/正常节目名.txt", "w")
file2 = open("../files/异常节目名.txt", "w")
# file3 = open("../files/全名.txt", "w")
# file4 = open("../files/全名-带括号.txt", "w")
file5 = open("../files/待比对节目名.txt", "w")
file6 = open("../files/结果-命中电影.txt", "w")
file7 = open("../files/结果-命中连续剧.txt", "w")
# 读取所有名称
for i in range(3):
    ws = wb.worksheets[i]
    nrows = ws.max_row  # 获得最大行数
    for j in range(2, nrows + 1):
        org_name.append(ws.cell(row=j, column=4).value)
# 按是否包含书名号拆成两个名称列表
for j in range(len(org_name)):
    if ("》" in org_name[j]) or ("《" in org_name[j]):
        out_name1.append(org_name[j])
    else:
        out_name2.append(org_name[j])
# 处理书名号，提取名称
for m in range(len(out_name1)):
    p1 = re.compile(r"[《](.*?)[》]", re.S)
    arr = re.findall(p1, out_name1[m])
    if arr:
        #                continue
        #                print (out_name1[m],arr)
        for i in arr:
            print(out_name1[m] + "  |  ", i, file=file1)
            print(i, file=file5)

    else:
        print(out_name1[m], file=file2)
# 处理全名的带括号
for n in range(len(out_name2)):
    p1 = re.compile(r"(.*?)[（]", re.S)
    arr = re.findall(p1, out_name2[n])
    if arr:
        #                continue
        #                print (out_name1[m],arr)
        print(out_name2[n] + "  |  ", arr[0], file=file1)
        print(arr[0], file=file5)

    else:
        print(out_name2[n] + "  |  ", out_name2[n], file=file1)
        print(out_name2[n], file=file5)

file1.close()
file2.close()
# file3.close
# file4.close
file5.close()
print("文件名提取完成")

# 查询数据库比对名称，采用精确匹配
with open("../files/待比对节目名.txt") as f:
    list1 = f.read().split("\n")
# print(list1)
print("比对开始，总共待比对节目: %d" % len(list1))


def sql_compare(list1, media_type, file):
    conn = cx_Oracle.connect("wacos/oss@172.18.10.10:1521/orcl")
    cursor = conn.cursor()
    print("连接数据库成功")
    print("开始比对%s" % media_type)
    nCount = 0
    for program_name in list1:
        now = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        nCount = nCount + 1
        if nCount % 100 == 0:
            print(now + " : Processed " + str(nCount) + " " + media_type)
        #    cursor = conn.cursor()
        sql_program = (
            (
                # trunk-ignore(flake8/E501)
                "select %sid,code,name,contentprovider,case status when '4' then '正常' when '9' then '已删除' else '其他' end,\
                case stockoutflag when '0' then '回收' when '1' then '上线' else '其他' end, createdate\
                    from %s where name=:1"
            )
            % (media_type, media_type)
        )
        sql_domainobjects = "select d.name,b.createdate from \
                (select domainid,createdate from domainobjects where objid=:1) b,domain d where b.domainid=d.domainid"
        sql_ctmshistoryeventdetail = "select d.name,b.createdate from \
                (select e.* from (select row_number() over(partition by c.domainid,c.objectaction order by c.underctmsnodeid ) rt,\
                    c.underctmsnodeid, c.domainid, c.objectaction,to_char(c.dispatchdetailcreatedate,'yyyy-mm-dd hh24:mi:ss') createdate from ctmshistoryeventdetail c \
                        where c.domainid is not null and c.objectaction='6' and c.objectid= :1)e where e.rt=1)b,\
                            domain d  where b.domainid=d.domainid"
        # 	print (sql)
        cursor.execute(sql_program, [program_name])
        program = cursor.fetchall()
        # 	if row:
        # 		print (row)
        for i in program:
            if i[5] == "上线":
                cursor.execute(sql_domainobjects, [i[0]])
                domainobjects = cursor.fetchall()
                for j in domainobjects:
                    print(i[1], i[2], i[3], j[0], i[5], j[1], sep=",", file=file)
            else:
                cursor.execute(sql_ctmshistoryeventdetail, [i[0]])
                ctmshistoryeventdetail = cursor.fetchall()
                for k in ctmshistoryeventdetail:
                    print(
                        i[1],
                        i[2],
                        i[3],
                        k[0],
                        i[5],
                        i[6],
                        k[1],
                        sep=",", 
                        file=file,
                    )

    #        print(i, file=file6)
    cursor.close()
    conn.close()
    now = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    print(now + " : Processed " + str(nCount) + " " + media_type)
    print("%s比对完成" % media_type)
    file.close()


sql_compare(list1, "program", file6)
sql_compare(list1, "series", file7)
