import datetime
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "故障1"

ws['A1'] = 42
ws.append([1, 2, 3])
ws['A2'] = datetime.datetime.now()
ws.append([34,56,7,656,3432,54])

wb.save("sample.xlsx")